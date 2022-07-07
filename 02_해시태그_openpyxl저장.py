from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import  pyautogui , time , os , re
import urllib.parse as rep
import urllib.request as req
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests as rq
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from typing import Optional,Union
from conf import get_login_info

class ChromeDriver:
    def __init__(self):
        # options 객체
        self.chrome_options = Options()

        # headless Chrome 선언
        self.chrome_options.add_argument('--headless')

        # 브라우저 꺼짐 방지
        self.chrome_options.add_experimental_option('detach', True)

        self.chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.104 Whale/3.13.131.36 Safari/537.36")

        # 불필요한 에러메시지 없애기
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        self.service = Service(executable_path=ChromeDriverManager().install())
        self.browser = webdriver.Chrome(service=self.service, options=self.chrome_options)
        self.browser.maximize_window()


class FaceBookLogin:
    def __init__(self):
        # ChromeDriver class의 browser 객체 불러오기
        self.browser = ChromeDriver().browser

        # 페이스북 id , pw 멤버변수로 정의
        self.id = get_login_info('FACEBOOK_ID')
        self.pw = get_login_info('FACEBOOK_PW')

    def login_execute(self)-> Union[None , int]:
        URL = "https://www.instagram.com/"

        print("""
        페이스북 로그인 중 ..""")
        # 페이지 상태 체크
        response = rq.get(URL)
        if response.status_code == 200:
            self.browser.get(URL)
            self.browser.implicitly_wait(5)

            try:
                # 페이스북으로 로그인하기 버튼 클릭
                fb_login = self.browser.find_element(By.CSS_SELECTOR, "button.sqdOP.yWX7d.y3zKF")
                self.browser.execute_script('arguments[0].click()', fb_login)
                time.sleep(1)

                # 페이스북 아이디 입력
                self.browser.find_element(By.CSS_SELECTOR, "input#email").send_keys(self.id)
                time.sleep(2)

                # 페이스북 패스워드 입력
                self.browser.find_element(By.CSS_SELECTOR, "input#pass").send_keys(self.pw)
                time.sleep(2)

                # 로그인 버튼 클릭
                login_btn = self.browser.find_element(By.CSS_SELECTOR, "button#loginbutton")
                self.browser.execute_script('arguments[0].click()', login_btn)
                time.sleep(2)

                # 페이지 잠깐 로딩
                WebDriverWait(self.browser, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))

                # 알림설정 문구 뜨는 경우 : 설정 누르기
                try:
                    # 페이지 잠깐 로딩
                    WebDriverWait(self.browser, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 'button.aOOlW.bIiDR')))

                    alarm_btn = self.browser.find_element(By.CSS_SELECTOR, "button.aOOlW.bIiDR")
                    self.browser.execute_script('arguments[0].click()', alarm_btn)
                    time.sleep(1)

                    print("""
                        --- 페이스북 로그인 완료 ! ---""")
                except:  # 알림설정 문구가 뜨지 않는경우
                    pass

                if "오류" in self.browser.find_element(By.CSS_SELECTOR, "body").text:
                    pyautogui.alert('페이지 오류!')
                    return 0

            except:  # 로그인 중 문제가 생긴 경우
                pyautogui.alert('로그인을 실패하였습니다!')
                return 0

        else:  # 페이지 status code가 200이 아닌경우
            pyautogui.alert(f"INSTAGRAM NOW PAGE STATUS CODE : {response.status_code} !!!")
            return 0

class AppInstagram:
    def __init__(self):
        # headers
        self.user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.57 Whale/3.14.133.23 Safari/537.36'
        self.headers = {'User-Agent': self.user_agent}

        # FaceBookLogin class 객체 멤버변수로 정의
        self.login = FaceBookLogin()

        # FaceBookLogin class에서 browser 객체 가져오기
        self.browser = self.login.browser

        # 해쉬태그 키워드 멤버변수 정의
        self.keword = self.hash_keword()


    def run(self)-> list:
        # return 데이터 담을 리스트변수 선언
        save_datas = []

        # 해시태그 페이지로 이동이 안되는 경우가 있음
        try:
            # 해시태그 페이지로 이동하기
            self.browser.get(URL)

            # 로딩 대기하기
            element = WebDriverWait(self.browser, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div._aagu'))
            )

            # 첫 번째 게시글 클릭하기
            a_tags = self.browser.find_elements(By.CSS_SELECTOR, 'div._aagu')[0]
            self.browser.execute_script('arguments[0].click()', a_tags)
            time.sleep(3)

            ## 수집할 게시물의 수 ##
            target_cnt = 500
            for i in range(1, target_cnt + 1):
                print(f"--- {i}개 ---  ")
                try:
                    save_datas.append(self.get_content(browser=self.browser, count=i))
                    self.move_next(browser=self.browser)
                except:
                    print("DATA CRAWLING FAILED")
                    time.sleep(2)
                    self.move_next(browser=self.browser)

            # 추출된 데이터들 리턴하기
            return save_datas

        # 해시태그 페이지로 이동이 안되는 경우 함수종료
        except:
            pyautogui.alert('현재 페이지에 문제가 있습니다!')
            return 0

    def hash_keword(self)-> str:
        while True :
            os.system('cls')
            keword = input("""원하시는 키워드명을 입력해주세요
    
    :""")
            if keword == "" :
                pyautogui.alert("키워드 값이 정상적이지 않습니다")
                os.system('cls')
                continue

            if re.match('[0-9]', keword):
                print("입력이 올바르지 않습니다")
                os.system('cls')
                continue

            return keword

    def move_next(self,browser)-> None :
        next_btn = browser.find_element(By.CSS_SELECTOR,'div._aaqg._aaqh > button._abl-')
        browser.execute_script('arguments[0].click()',next_btn)
        time.sleep(2)

    def get_content(self,browser , count)-> list:
        # 댓글 더보기 버튼 클릭
        try :
            loop_cnt = 0
            while True :
                if loop_cnt == 1 :
                    break
                more_btn = browser.find_element(By.CSS_SELECTOR,'div._ab8w._ab94._ab99._ab9h._ab9m._ab9p._abcj > button')
                browser.execute_script('arguments[0].click()',more_btn)
                time.sleep(2)

                # 루프 카운트 증감
                loop_cnt += 1
        except :
            pass

        # 추출데이터 담을 리스트 변수 선언
        save_data = []

        # 현재 드라이버의 페이지소스 파싱하기
        soup = bs(browser.page_source,'html.parser')

        # 인스타그램 닉네임
        insta_nickname = soup.select('span._aap6._aap7._aap8 > a')[0]

        # 본문내용
        content = soup.select('div._a9zs > span')[0]

        # 댓글내용 -> List Comprehension 문법으로 작성
        comment: Union[list,str] = [x.text.strip() for x in soup.select('div._a9zs > span') if x != content]

        # 댓글 개수
        comment_cnt = len(comment)

        # 댓글내용 전처리
        comment = re.sub('[\[\]]','',str(comment))

        # 게시글 등록일
        content_time = soup.select('time._a9ze._a9zf')[0]

        # 좋아요 개수
        like_cnt = soup.select_one('div._aacl._aaco._aacw._aacx._aada._aade > span')

        # 이미지 가져오기 (여러 이미지가 있는 경우 썸네일만 가져옴)
        image = soup.select_one('div.ll8tlv6m.rq0escxv.j83agx80.taijpn5t.tgvbjcpo.hpfvmrgz.hzruof5a img._aagt')

        # content 분기처리
        if content == None or content.text == '':
            content = '-'
        else :
            content = content.text.strip()
            content = re.sub('#.*','',content)

        # content_time 분기처리
        if content_time == None or content_time.text == '':
            content_time = '-'
        else :
            content_time = content_time.text.strip()

        # 인스타 닉네임 분기처리
        if insta_nickname == None or insta_nickname.text == '':
            insta_nickname = 'None'
        else :
            insta_nickname = insta_nickname.text.strip()

        if like_cnt == None or like_cnt.text == '':
            like_cnt = 0
        else :
            like_cnt = like_cnt.text.strip()
            like_cnt = int(re.sub('[,]','',like_cnt))

        # 이미지 링크주소 추출하기
        image_link = image.attrs['src']

        # 이미지 다운로드
        img = self.download_image(img_lnk=image_link,count=count)

        # save_data 변수에 추출데이터 담기
        save_data.append([img,insta_nickname,like_cnt,content_time,comment_cnt,content,comment])

        # 추출데이터 출력하기
        print(f"인스타그램 아이디 : {insta_nickname}\n좋아요 수 : {like_cnt}\n게시일자 : {content_time}\n댓글 수 : {comment_cnt}\n본문 내용 : {content}\n댓글 내용 : {comment}\n")

        # 추출데이터 리턴하기
        return save_data

    def download_image(self,img_lnk,count):
        img_path = os.path.abspath('img')
        img_filename = os.path.join(img_path,f'{count}.png')

        if not os.path.exists(img_path) :
            os.mkdir(img_path)

        # 이미지 다운로드하기
        req.urlretrieve(img_lnk,img_filename)

        # 추출된 이미지파일 변수에 넣기
        img = Image(img_filename)
        img.height = 20
        img.width  = 70

        return img

class OpenPyXL:
    def __init__(self):
        # AppInstagram 객체 멤버변수로 정의
        self.app = AppInstagram()

        # run 메소드 실행 후 , return 값 멤버변수로 정의
        self.results = self.app.run()

        if self.results != 0 :
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = f"{self.app.keword}"
            self.ws.append(["Thumbnail", "INSTAGRAM ID", "LIKE COUNT", "POSTED DATE", "COMMENT COUNT", "CONTENT", "COMMENT"])

            # 데이터 저장할 경로 지정하기
            self.savePath = os.path.abspath(f'인스타그램_{self.app.keword}')
            self.fileName = f"{self.app.keword}.xlsx"

            # 파일 저장하기
            self.savefile()

    def savefile(self)-> None:
        row = 2
        for x in self.results:
            for data in x:
                self.ws.add_image(data[0], f"A{row}")
                self.ws[f"B{row}"] = data[1]
                self.ws[f"C{row}"] = data[2]
                self.ws[f"D{row}"] = data[3]
                self.ws[f"E{row}"] = data[4]
                self.ws[f"F{row}"] = data[5]
                self.ws[f"G{row}"] = data[6]

                row += 1
        savePath = os.path.abspath('인스타그램_해쉬태그')
        fileName = 'result.xlsx'

        if not os.path.exists(savePath):
            os.mkdir(savePath)
        self.wb.save(os.path.join(savePath,fileName))
        self.wb.close()

        pyautogui.alert(f'파일 저장완료{savePath}')



if __name__ == '__main__' :
    app = AppInstagram()

    app.run()







